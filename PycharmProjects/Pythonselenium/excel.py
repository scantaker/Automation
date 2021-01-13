import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Color, PatternFill
from openpyxl.styles import colors

from openfile import file_paths_openfile
from get_ave_accuracy_by_column import get_ave_accuracy_by_column

file_paths = file_paths_openfile()

ave_doc_performance_file_allfiles = 0


#calculate same cell value ave from multi files
for file_path in file_paths:
    try:
        wbload = load_workbook(file_path, data_only=True)

    except IOError:
        print("File can't open")

    KPISheet = wbload.active

    ave_doc_performance_file_singlefile = get_ave_accuracy_by_column(0.04, KPISheet, 4)
    print("current file doc perf:", ave_doc_performance_file_singlefile)
    ave_doc_performance_file_allfiles += ave_doc_performance_file_singlefile

ave_ave_doc_performance_file_allfiles = ave_doc_performance_file_allfiles / len(file_paths)
print("ave file doc perf:", ave_ave_doc_performance_file_allfiles)


#save file
saved_wb = openpyxl.Workbook()
saved_ws = saved_wb.active

saved_ws.title = "aveperformance"
saved_ws['A1'] = "Average of doc performance for all docs"
A1 = saved_ws['A1']

##set font
A1.font = Font(color = colors.DARKBLUE, bold=True, size=24)
A1.fill = PatternFill(fgColor=colors.GREEN, bgColor=colors.YELLOW)

saved_ws ['A2'] = ave_ave_doc_performance_file_allfiles

saved_wb.save("aveperformance.xlsx")

