import openpyxl
from openpyxl import Workbook


def get_ave_accuracy_by_column(switch, KPISheet, column_name):
    """
    :param switch:      User input number <0.04 to get all noise level <0.04, or >0.04
    :param KPISheet:    Active sheet that user open from workbook
    :param column_name: Number of column, eg: column A=1, Column B=2
    :return:            Average the number from column_name, for all open files
    """
    count_less4 = 0
    count_greater4 = 0

    sum_allless4 = 0
    sum_all_greater4 = 0

    for noiseRow in range(1, KPISheet.max_row):
        noise_level_split = KPISheet.cell(row=noiseRow, column=3).value

        if type(noise_level_split) is float or type(noise_level_split) is int:
            if noise_level_split < 0.04:
                count_less4 += 1
                sum_allless4 += KPISheet.cell(row=noiseRow, column=column_name).value

            if noise_level_split >= 0.04:
                count_greater4 += 1
                sum_all_greater4 += KPISheet.cell(row=noiseRow, column=column_name).value

    assert (count_less4 != 0)
    assert (count_greater4 != 0)
    avg_all_less4 = sum_allless4 / count_less4
    avg_all_greater4 = sum_all_greater4 / count_greater4

    if switch < 0.04:
        print("current file count of noise level <4%", count_less4)
        return avg_all_less4

    else:
        print("current file count of noise level >=4%", count_greater4)
        return avg_all_greater4

