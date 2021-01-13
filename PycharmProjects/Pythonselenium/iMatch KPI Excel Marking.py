from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Color, PatternFill
from openpyxl.styles import colors
import re
from tkinter import messagebox

from openfile import file_paths_openfile


def mark_valid_datapoint():
    """
    If it's valid datapoint, mark 1, otherwise 0
    :return: count_changed_rows, how many rows change in each file
    """
    count_changed_rows = 0

    for datapoint_row in range(1, KPISheet.max_row+1):
        if KPISheet.cell(row=datapoint_row, column=2).value is not None or KPISheet.cell(row=datapoint_row, column=3).value is None \
                or KPISheet.cell(row=datapoint_row, column=5).value is None or KPISheet.cell(row=datapoint_row, column=4).value is None:
            continue

        if KPISheet.cell(row=datapoint_row, column=4).value == "NA":
            KPISheet.cell(row=datapoint_row, column=8).value = 0
            KPISheet.cell(row=datapoint_row, column=9).value = 0
            KPISheet.cell(row=datapoint_row, column=10).value = 0
            KPISheet.cell(row=datapoint_row, column=11).value = 0
            KPISheet.cell(row=datapoint_row, column=4).fill = PatternFill(fgColor=colors.YELLOW, fill_type="solid")
            KPISheet.cell(row=datapoint_row, column=5).fill = PatternFill(fgColor=colors.YELLOW, fill_type="solid")

        elif KPISheet.cell(row=datapoint_row, column=5).value == KPISheet.cell(row=datapoint_row, column=4).value:
            KPISheet.cell(row=datapoint_row, column=8).value = 1
            KPISheet.cell(row=datapoint_row, column=9).value = 0
            KPISheet.cell(row=datapoint_row, column=10).value = 0
            KPISheet.cell(row=datapoint_row, column=11).value = 1

        elif str(KPISheet.cell(row=datapoint_row, column=5).value) in str(KPISheet.cell(row=datapoint_row, column=4).value) or \
            str(KPISheet.cell(row=datapoint_row, column=4).value) in str(KPISheet.cell(row=datapoint_row, column=5).value):
            KPISheet.cell(row=datapoint_row, column=8).value = 1
            KPISheet.cell(row=datapoint_row, column=9).value = 1
            KPISheet.cell(row=datapoint_row, column=10).value = 0
            KPISheet.cell(row=datapoint_row, column=11).value = 0

            KPISheet.cell(row=datapoint_row, column=4).font = Font(color=colors.BLUE)
            KPISheet.cell(row=datapoint_row, column=5).font = Font(color=colors.BLUE)

        else:
            KPISheet.cell(row=datapoint_row, column=8).value = 1
            KPISheet.cell(row=datapoint_row, column=9).value = 0
            KPISheet.cell(row=datapoint_row, column=10).value = 0
            KPISheet.cell(row=datapoint_row, column=11).value = 0

            ##set font
            KPISheet.cell(row=datapoint_row, column=4).font = Font(color=colors.RED)
            KPISheet.cell(row=datapoint_row, column=5).font = Font(color=colors.RED)

        count_changed_rows += 1
    return count_changed_rows


file_paths = file_paths_openfile()

for file_path in file_paths:
    ##TO GET THE FILE NAME, LATER USE FOR NEW FILENAME
    file_path_names = re.split('/', file_path)
    file_path_name = file_path_names [len(file_path_names)-1]

    try:
        wbload = load_workbook(file_path, data_only=True)

    except IOError as e:
        messagebox.showerror("ERROR file "+file_path_name, "can't open!")

    ###Ensure the 2nd excel tab is used
    sheet_names = wbload.sheetnames
    KPISheet = wbload.get_sheet_by_name(sheet_names[1])
    print(sheet_names[1])

    ###warning if excel file is not valid
    if mark_valid_datapoint() == 0:
        messagebox.showerror("ERROR in file "+file_path_name, "No valid data or wrong format or you have hidden sheet!")
        continue

    save_file_name = "NEW " + str(file_path_name)
    wbload.save(save_file_name)
